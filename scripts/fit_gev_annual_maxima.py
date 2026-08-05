"""
Fit GEV to annual-maximum daily rainfall (Nepal, 280 stations, 1980-2020).

Method note (2026-08-04 session):
  GEV is fit to block maxima, not the raw daily series. The daily zero-atom
  (65-80% dry days per station) lives in the daily data and does not reach
  the fit once annual maxima are extracted -- an annual max is zero only if
  every day that year was dry, which does not happen in a monsoon climate.
  The real decision at this step is the completeness rule below: a year
  with too many missing days gives a biased-low "maximum" and should be
  dropped rather than fit.

Source: data/Daily Rainfall all nepal.xlsx, sheet "Sheet4".
  Row 3 = station code, row 4 = station name, row 5+ = date, then one
  column of daily rainfall (mm) per station.
"""
import json
from collections import defaultdict

import numpy as np
import openpyxl
from scipy import stats

SOURCE = "data/Daily Rainfall all nepal.xlsx"
COMPLETENESS_THRESHOLD = 0.90  # fraction of days in a year required to keep that year's max

# Station codes selected for the 2026-08-04 session: the 5 stations with a
# 100%-complete 1980-2020 daily record (found by scanning all 280 columns).
TARGET_CODES = [1030, 804, 1319, 406, 909]  # Kathmandu, Pokhara, Biratnagar, Surkhet, Simara


def is_leap(year: int) -> bool:
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def extract_annual_maxima(ws, target_codes, completeness=COMPLETENESS_THRESHOLD):
    codes = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    names = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))

    target_cols = {code: codes.index(code) for code in target_codes}
    col_to_code = {v: k for k, v in target_cols.items()}
    cols = set(target_cols.values())

    year_count = {c: defaultdict(int) for c in cols}
    year_max = {c: defaultdict(lambda: -1.0) for c in cols}

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        date = row[0]
        if date is None:
            continue
        yr = date.year
        for c in cols:
            v = row[c]
            if v is None or v == "":
                continue
            try:
                v = float(v)
            except (ValueError, TypeError):
                continue
            year_count[c][yr] += 1
            if v > year_max[c][yr]:
                year_max[c][yr] = v

    results = {}
    for c in cols:
        code = col_to_code[c]
        name = str(names[c])
        ams = []
        for yr in sorted(year_count[c].keys()):
            days = 366 if is_leap(yr) else 365
            if year_count[c][yr] >= completeness * days:
                ams.append(year_max[c][yr])
        results[code] = {"name": name, "ams": np.array(ams)}
    return results


def fit_gev(ams: np.ndarray):
    """scipy uses shape convention c = -xi; return (xi, mu, sigma)."""
    c_fit, loc, scale = stats.genextreme.fit(ams)
    return -c_fit, loc, scale


def bootstrap_ci_xi(xi, mu, sigma, n, n_boot=150, seed=42):
    """Parametric bootstrap CI for xi -- expect it to be wide/asymmetric at n~40."""
    rng = np.random.default_rng(seed)
    c_fit = -xi
    boots = []
    for _ in range(n_boot):
        sample = stats.genextreme.rvs(c_fit, loc=mu, scale=sigma, size=n, random_state=rng)
        try:
            cb, _, _ = stats.genextreme.fit(sample)
            boots.append(-cb)
        except Exception:
            continue
    boots = np.array(boots)
    return boots.std(), np.percentile(boots, 2.5), np.percentile(boots, 97.5)


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Sheet4"]

    extracted = extract_annual_maxima(ws, TARGET_CODES)

    print(f"{'Station':22s} {'code':>5s} {'n_yrs':>6s} {'xi':>8s} {'mu':>8s} {'sigma':>8s}")
    out = {}
    for code, d in extracted.items():
        ams = d["ams"]
        xi, mu, sigma = fit_gev(ams)
        out[code] = {"name": d["name"], "n": len(ams), "xi": xi, "mu": mu, "sigma": sigma,
                      "ams": ams.tolist()}
        print(f"{d['name']:22s} {code:>5d} {len(ams):>6d} {xi:>8.3f} {mu:>8.2f} {sigma:>8.2f}")

    with open("outputs/gev_annual_maxima_results.json", "w") as f:
        json.dump(out, f, indent=2)
